#!/usr/bin/env python3
"""
Multi-Bank PDF Statement Parser
Converts HDFC/SBI/IOB/Canara/Indian Bank statement PDFs to CSV/Excel using coordinate-based layout parsing.
Author: Antigravity AI
"""

from __future__ import annotations

import os
import sys
import re
import csv
import argparse
import warnings
from datetime import datetime
from dataclasses import dataclass, field, replace

# Module-level compiled date sub-patterns
MONTH_PATTERN = re.compile(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)$', re.IGNORECASE)
YEAR_PATTERN = re.compile(r'^\d{4}$')

# Tolerant combined "dd Mon yyyy" matcher for split-date layouts (Indian Bank):
# joins all words in the date column first, so tokenizer differences (e.g. "05 Apr"
# extracted as one word, single-digit days, trailing period) don't break anchoring.
SPLIT_DATE_PATTERN = re.compile(r'^(\d{1,2})\s+([A-Za-z]{3})\.?\s+(\d{4})$')

# Try to import fitz (PyMuPDF) or pdfplumber
HAS_FITZ = False
try:
    import fitz
    HAS_FITZ = True
except ImportError:
    pass

HAS_PDFPLUMBER = False
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    pass

# Try to import openpyxl for Excel output
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class NormalizedWord:
    """Standardized word wrapper for both PyMuPDF and pdfplumber."""
    __slots__ = ('x0', 'top', 'x1', 'text')

    def __init__(self, x0, top, x1, text):
        self.x0 = float(x0)
        self.top = float(top)
        self.x1 = float(x1)
        self.text = str(text)


@dataclass
class BankProfile:
    """Declarative specification for bank-specific PDF layouts."""
    name: str
    display_name: str
    date_pattern: re.Pattern
    date_x_range: tuple[float, float]
    col_bounds: dict[str, tuple[float, float] | None]
    page1_min_y: float
    pageN_min_y: float
    max_y: float = 790.0
    date_type: str = "single"  # "single", "split_3", "iob"
    footer_keywords: list[str] = field(default_factory=list)
    header_keywords: list[str] = field(default_factory=list)
    chronological: bool = True
    line_group_tolerance: float = 2.0
    debit_dash_x_range: tuple[float, float] | None = None
    credit_dash_x_range: tuple[float, float] | None = None
    currency_prefix: str | None = None
    vdate_pattern: re.Pattern | None = None
    iob_block_offset_before: float = 6.0
    iob_block_offset_after: float = 15.0
    adaptive_bounds: bool = False  # derive column bounds from page-1 header row


BANK_PROFILES = {
    'hdfc': BankProfile(
        name='hdfc',
        display_name='HDFC Bank',
        date_pattern=re.compile(r'^\d{2}/\d{2}/\d{2,4}$'),
        date_x_range=(30.0, 65.0),
        col_bounds={
            'date': (30.0, 65.0),
            'narration': (65.0, 280.0),
            'chq_ref': (280.0, 350.0),
            'val_date': (350.0, 400.0),
            'withdrawal': (400.0, 480.0),
            'deposit': (480.0, 560.0),
            'balance': (560.0, 630.0),
        },
        page1_min_y=225.0,
        pageN_min_y=40.0,  # Fixed: Page 2+ starts table at top
        max_y=780.0,
        footer_keywords=['STATEMENT SUMMARY', 'TOTAL', 'CLOSING BALANCE'],
        header_keywords=['Date', 'Narration', 'Chq/Ref']
    ),
    'sbi': BankProfile(
        name='sbi',
        display_name='State Bank of India (SBI)',
        date_pattern=re.compile(r'^\d{2}/\d{2}/\d{4}$'),
        date_x_range=(20.0, 78.0),
        col_bounds={
            'date': (20.0, 78.0),
            'val_date': (78.0, 130.0),
            'narration': (130.0, 290.0),
            'chq_ref': (290.0, 335.0),
            'withdrawal': (335.0, 410.0),
            'deposit': (410.0, 485.0),
            'balance': (485.0, 580.0),
        },
        page1_min_y=180.0,
        pageN_min_y=40.0,
        max_y=730.0,
        footer_keywords=['Statement Summary', 'Total:', 'Closing Balance', 'Page ', 'This is a computer'],
        header_keywords=['Txn Date', 'Value Date', 'Description', 'Ref No']
    ),
    'iob': BankProfile(
        name='iob',
        display_name='Indian Overseas Bank (IOB)',
        date_pattern=re.compile(r'^\d{2}-[A-Z][a-z]{2}-\d{2}$'),
        vdate_pattern=re.compile(r'^\((\d{2}-[A-Z][a-z]{2}-\d{2})\)$'),
        date_x_range=(40.0, 90.0),
        col_bounds={
            'date': (40.0, 90.0),
            'narration': (90.0, 275.0),
            'chq_ref': (275.0, 340.0),
            'withdrawal': (395.0, 455.0),
            'deposit': (455.0, 510.0),
            'balance': (510.0, 600.0),
        },
        page1_min_y=260.0,
        pageN_min_y=30.0,
        max_y=790.0,
        date_type='iob',
        footer_keywords=['STATEMENT OF THE ACCOUNT', 'CUSTOMER DETAILS', 'Effective available balance', 'computer generated statement', 'Page '],
        header_keywords=['Particulars', 'Ref No.', 'Debit(Rs)'],
        chronological=False,
        iob_block_offset_before=6.0,
        iob_block_offset_after=15.0
    ),
    'canara': BankProfile(
        name='canara',
        display_name='Canara Bank',
        date_pattern=re.compile(r'^\d{2}-[A-Z][a-z]{2}-\d{2}$'),
        date_x_range=(30.0, 95.0),
        col_bounds={
            'date': (30.0, 95.0),
            'narration': (90.0, 310.0),
            'chq_ref': None,
            'val_date': None,
            'withdrawal': (310.0, 424.0),
            'deposit': (424.0, 515.0),
            'balance': (515.0, 650.0),
        },
        page1_min_y=240.0,
        pageN_min_y=10.0,
        max_y=840.0,
        footer_keywords=['Total:', 'Opening Balance', 'Closing Balance', 'Dr Count', 'Cr Count', 'UNLESS THE', 'COMPUTER OUTPUT', 'End of Statement'],
        header_keywords=['Txn Date', 'Txn Description', 'Debit', 'Credit', 'Balance']
    ),
    'indianbank': BankProfile(
        name='indianbank',
        display_name='Indian Bank',
        date_pattern=re.compile(r'^\d{2}$'),
        date_x_range=(71.0, 124.0),
        col_bounds={
            'date': (71.0, 124.0),
            'narration': (145.0, 260.0),
            'chq_ref': None,
            'val_date': None,
            'withdrawal': (283.0, 330.0),
            'deposit': (378.0, 425.0),
            'balance': (470.0, 525.0),
        },
        page1_min_y=65.0,
        pageN_min_y=65.0,
        max_y=780.0,
        date_type='split_3',
        debit_dash_x_range=(305.0, 308.0),
        credit_dash_x_range=(400.0, 403.0),
        currency_prefix='INR',
        adaptive_bounds=True
    )
}


def _matches_word_keyword(keyword: str, line_text: str) -> bool:
    """Helper to check if a keyword matches in line text without false substring triggers or special character issues."""
    if not keyword or not line_text:
        return False
    # If keyword has trailing space (e.g. 'Page '), match line prefix
    if keyword.endswith(' '):
        return line_text.strip().lower().startswith(keyword.strip().lower() + ' ')
    kw_clean = keyword.strip()
    pattern = r'(?:^|\s)' + re.escape(kw_clean) + r'(?:\s|$)'
    return bool(re.search(pattern, line_text, re.IGNORECASE))


class BankStatementParser:
    """Unified configuration-driven parser for HDFC, SBI, IOB, Canara, and Indian Bank statement PDFs."""

    def __init__(self, pdf_path, verbose=False, bank=None):
        if not HAS_FITZ and not HAS_PDFPLUMBER:
            raise ImportError("Neither 'pymupdf' nor 'pdfplumber' is installed. Please install at least one (e.g. 'pip install pymupdf').")
        self.pdf_path = pdf_path
        self.verbose = verbose
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        self.bank = (bank or self.detect_bank(pdf_path)).lower()

    # ── Bank auto-detection ──────────────────────────────────────────────

    @staticmethod
    def detect_bank(pdf_path):
        """Read page 1 text and detect the bank using precise pattern matching."""
        try:
            if HAS_FITZ:
                doc = fitz.open(pdf_path)
                text = doc[0].get_text()
                doc.close()
            elif HAS_PDFPLUMBER:
                with pdfplumber.open(pdf_path) as pdf:
                    text = pdf.pages[0].extract_text() or ""
            else:
                text = ""

            text_lower = text.lower()
            if "state bank of india" in text_lower:
                return "sbi"
            if "indian overseas bank" in text_lower or "ioba" in text_lower:
                return "iob"
            if "canara bank" in text_lower or "e-pass sheet" in text_lower:
                return "canara"
            if "indian overseas bank" not in text_lower and ("indian bank" in text_lower or "indianbank" in text_lower or "idib" in text_lower):
                return "indianbank"
            if "hdfc bank" in text_lower:
                return "hdfc"

            warnings.warn("Could not auto-detect bank from PDF header. Defaulting to 'hdfc'. Pass --bank to override explicitly.")
            return "hdfc"
        except Exception as e:
            warnings.warn(f"Error during bank auto-detection ({e}). Defaulting to 'hdfc'. Pass --bank to override explicitly.")
            return "hdfc"

    # ── Shared helpers ──────────────────────────────────────────────────

    def clean_amount(self, amount_str):
        """Cleans formatting from amount strings and converts to float representation."""
        if not amount_str:
            return None
        cleaned = str(amount_str).replace(",", "").strip()
        if not cleaned or cleaned in ('-', '(', ')'):
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None

    def format_amount(self, val):
        if val is None:
            return ""
        return f"{val:.2f}"

    def _extract_iob_dates_and_ref(self, block_ys, grouped_lines, profile):
        """Helper to extract transaction dates, value dates, and reference numbers for IOB blocks."""
        date_text = ""
        val_date_text = ""
        ref_parts = []

        ref_bounds = profile.col_bounds.get('chq_ref')

        for gk in block_ys:
            line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)
            for w in line_words:
                if profile.date_x_range[0] <= w.x0 <= profile.date_x_range[1]:
                    if profile.vdate_pattern:
                        vm = profile.vdate_pattern.match(w.text)
                        if vm:
                            val_date_text = vm.group(1)
                    if profile.date_pattern.match(w.text) and not date_text:
                        date_text = w.text

            if ref_bounds:
                rx0, rx1 = ref_bounds
                r_words = [w.text for w in line_words if rx0 <= w.x0 < rx1 and w.text != '-']
                if r_words:
                    ref_parts.extend(r_words)

        return date_text, val_date_text or date_text, " ".join(ref_parts).strip()

    def _handle_cross_page_continuation(self, raw_txs, valid_ys, date_ys, grouped_lines, profile):
        """Appends narration overflow at top of page (before first anchor) to previous page's final transaction."""
        if not raw_txs or not valid_ys:
            return
        first_date_y = date_ys[0][0] if date_ys else (valid_ys[-1] + 1)
        pre_date_ys = [gk for gk in valid_ys if gk < first_date_y]
        if pre_date_ys:
            narration_bounds = profile.col_bounds.get('narration')
            if narration_bounds:
                nx0, nx1 = narration_bounds
                pre_parts = []
                for gk in pre_date_ys:
                    line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)
                    n_words = [w.text for w in line_words if nx0 <= w.x0 < nx1]
                    if n_words:
                        pre_parts.append(" ".join(n_words))
                if pre_parts:
                    raw_txs[-1]['narration'] += " " + " ".join(pre_parts)

    def _infer_split_amounts(self, raw_txs, balance_val, debit_val, credit_val):
        """Balance-chain inference for split-date profiles (Indian Bank).

        Uses the running balance delta to validate and, if needed, repair the
        withdrawal/deposit assignment. Parsed values (dash indicators + column
        extraction) act as hints; the mathematically implied amounts win on any
        conflict. Returns the (withdrawal, deposit) pair to store.
        """
        if not raw_txs:
            # First transaction of the statement: no previous balance to infer from.
            return debit_val, credit_val

        prev_balance = raw_txs[-1]['balance']
        delta = round(balance_val - prev_balance, 2)

        if abs(delta) < 0.005:
            exp_debit, exp_credit = None, None
        elif delta < 0:
            exp_debit, exp_credit = round(-delta, 2), None
        else:
            exp_debit, exp_credit = None, round(delta, 2)

        def _matches(parsed, expected):
            if expected is None:
                return parsed is None
            return parsed is not None and abs(parsed - expected) <= 0.02

        if _matches(debit_val, exp_debit) and _matches(credit_val, exp_credit):
            return debit_val, credit_val

        print(f"  Amount inference: repaired W/D from balance chain (delta {delta:+.2f}; "
              f"parsed W={debit_val}, D={credit_val} -> W={exp_debit}, D={exp_credit})")
        return exp_debit, exp_credit

    def _infer_chain_amounts(self, txs):
        """Balance-chain validation/repair pass over a CHRONOLOGICAL transaction list.

        Used for reverse-chronological profiles (IOB) after re-sorting: the running
        balance delta mathematically determines withdrawal vs deposit, so parsed
        values that disagree with the chain are corrected (loudly).
        """
        repaired = 0

        def _matches(parsed, expected):
            if expected is None:
                return parsed is None
            return parsed is not None and abs(parsed - expected) <= 0.02

        for i in range(1, len(txs)):
            prev_bal = txs[i - 1]['balance']
            bal = txs[i]['balance']
            delta = round(bal - prev_bal, 2)

            if abs(delta) < 0.005:
                exp_w, exp_d = None, None
            elif delta < 0:
                exp_w, exp_d = round(-delta, 2), None
            else:
                exp_w, exp_d = None, round(delta, 2)

            w, d = txs[i]['withdrawal'], txs[i]['deposit']
            if _matches(w, exp_w) and _matches(d, exp_d):
                continue

            repaired += 1
            if repaired <= 10:
                print(f"  Amount inference: repaired W/D from balance chain (delta {delta:+.2f}; "
                      f"parsed W={w}, D={d} -> W={exp_w}, D={exp_d})")
            txs[i]['withdrawal'], txs[i]['deposit'] = exp_w, exp_d

        if repaired:
            print(f"  Amount inference: {repaired} transaction(s) repaired via balance chain.")
        return txs

    def _parse_amount_text(self, text):
        """Normalizes an amount token: strips currency prefixes/symbols (INR, Rs, \u20b9)
        and trailing Dr/Cr suffixes, handles parenthesized negatives, then cleans."""
        if not text:
            return None
        t = str(text).strip()
        t = re.sub(r'(?i)\b(?:INR|RS)\b\.?', '', t)
        t = t.replace('\u20b9', '')
        neg = t.startswith('(') and t.endswith(')')
        if neg:
            t = t[1:-1]
        t = re.sub(r'(?i)(?:DR|CR)\.?$', '', t).strip()
        val = self.clean_amount(t)
        if val is not None and neg:
            val = -val
        return val

    def _first_amount_in_range(self, line_words, x0, x1):
        """Returns the first parseable amount among words whose x0 falls within [x0, x1].

        Unlike a plain first-word grab, tokens that fail a naive float parse (e.g.
        'INR', '1,234.56Dr') don't abort the scan \u2014 subsequent words are still tried.
        """
        for w in line_words:
            if x0 <= w.x0 <= x1:
                val = self._parse_amount_text(w.text)
                if val is not None:
                    return val
        return None

    def _apply_adaptive_bounds(self, words, profile):
        """Derives column x-bounds from the page-1 header row instead of static points.

        Locates the header line (the one mentioning several column labels), computes
        each column's center from its label position, and places boundaries at the
        midpoints between adjacent centers. Falls back to the profile's static bounds
        when the header cannot be reliably recognized or the geometry looks wrong.
        """
        grouped = {}
        for w in words:
            placed = False
            for gk in grouped.keys():
                if abs(gk - w.top) < profile.line_group_tolerance:
                    grouped[gk].append(w)
                    placed = True
                    break
            if not placed:
                grouped[w.top] = [w]

        classifiers = (
            (('date',), 'date'),
            (('descr', 'narration', 'particul', 'transaction', 'details'), 'narration'),
            (('debit', 'withdrawal'), 'withdrawal'),
            (('credit', 'deposit'), 'deposit'),
            (('balance',), 'balance'),
        )
        order = ['date', 'narration', 'withdrawal', 'deposit', 'balance']

        header_centers = None
        for gk in sorted(grouped.keys()):
            line_words = sorted(grouped[gk], key=lambda w: w.x0)
            col_words = {}
            for w in line_words:
                lt = w.text.lower()
                for needles, col in classifiers:
                    if any(n in lt for n in needles):
                        col_words.setdefault(col, []).append(w)
                        break
            if not all(c in col_words for c in ('withdrawal', 'deposit', 'balance')):
                continue
            centers = {}
            for col, ws in col_words.items():
                ws = sorted(ws, key=lambda w: w.x0)
                # Merge adjacent label words into one extent (e.g. 'Transaction' + 'Details');
                # keep only the leftmost contiguous group (e.g. 'Txn Date' vs 'Value Date').
                group = [ws[0]]
                for w in ws[1:]:
                    if w.x0 - group[-1].x1 < 10.0:
                        group.append(w)
                    else:
                        break
                centers[col] = (group[0].x0 + group[-1].x1) / 2
            header_centers = centers
            break

        if header_centers is None:
            print("Adaptive bounds: header row not recognized; keeping static column bounds.")
            return profile

        present = [c for c in order if c in header_centers]
        cs = [header_centers[c] for c in present]
        # Sanity: centers must be strictly increasing with meaningful spacing
        if any(cs[i + 1] - cs[i] < 20.0 for i in range(len(cs) - 1)):
            print("Adaptive bounds: implausible header geometry; keeping static column bounds.")
            return profile

        new_bounds = dict(profile.col_bounds)
        n = len(present)
        for i, col in enumerate(present):
            left = (cs[i - 1] + cs[i]) / 2 if i > 0 else max(0.0, cs[0] - 45.0)
            right = (cs[i] + cs[i + 1]) / 2 if i < n - 1 else cs[-1] + 60.0
            new_bounds[col] = (left, right)

        # Safety net: every derived bound must overlap the known-good static bound,
        # otherwise the header was misread and we fall back to static bounds entirely.
        for col in present:
            static = profile.col_bounds.get(col)
            if static:
                d0, d1 = new_bounds[col]
                if d1 < static[0] or d0 > static[1]:
                    print(f"Adaptive bounds: derived '{col}' ({d0:.0f}-{d1:.0f}) doesn't overlap "
                          f"static ({static[0]:.0f}-{static[1]:.0f}); keeping static column bounds.")
                    return profile

        print("Adaptive bounds derived from page-1 header: "
              + ", ".join(f"{c}={new_bounds[c][0]:.0f}-{new_bounds[c][1]:.0f}" for c in present))
        return replace(profile, col_bounds=new_bounds)

    # ── Universal Unified Parsing Engine ─────────────────────────────────

    def parse(self):
        profile = BANK_PROFILES.get(self.bank, BANK_PROFILES['hdfc'])
        print(f"Detected bank: {profile.display_name}")

        if HAS_FITZ:
            try:
                return self._parse_with_engine(profile, engine="fitz")
            except Exception as e:
                print(f"PyMuPDF parsing failed: {e}. Falling back to pdfplumber...")

        return self._parse_with_engine(profile, engine="pdfplumber")

    def _extract_words(self, page, engine):
        """Extracts words from a PDF page and normalizes them into NormalizedWord objects."""
        words = []
        if engine == "fitz":
            raw_words = page.get_text("words")
            for w in raw_words:
                words.append(NormalizedWord(w[0], w[1], w[2], w[4]))
        else:
            raw_words = page.extract_words()
            for w in raw_words:
                words.append(NormalizedWord(w['x0'], w['top'], w['x1'], w['text']))
        return words

    def _parse_with_engine(self, profile: BankProfile, engine: str):
        """
        Generic, Profile-Driven Bank PDF Parsing Engine.

        Flow:
        1. Extract & normalize words via PyMuPDF (fitz) or pdfplumber.
        2. Filter header padding (page1_min_y vs pageN_min_y) and footer limits (max_y).
        3. Group words into horizontal lines by Y-coordinate (configurable tolerance).
        4. Suppress header rows and footer summary blocks using word-boundary keyword triggers.
        5. Identify date anchors (single-pattern, split 3-word date, or balance-anchored for IOB).
        6. Segment horizontal lines into discrete transaction blocks between consecutive anchors.
        7. Support cross-page narration overflow appending to previous page's final transaction across all bank profiles.
        8. Extract fields (date, val_date, narration, chq_ref, withdrawal, deposit, balance).
        9. Standardize & validate numerical values, handle dash indicators & currency prefixes.
        10. Re-order pages and lines if profile specifies non-chronological order (e.g. IOB).
        """
        print(f"Opening PDF ({'PyMuPDF' if engine == 'fitz' else 'pdfplumber'}): {self.pdf_path}")
        raw_txs = []

        if engine == "fitz":
            doc = fitz.open(self.pdf_path)
            total_pages = len(doc)
            get_page = lambda idx: doc[idx]
            close_doc = lambda: doc.close()
        else:
            pdf = pdfplumber.open(self.pdf_path)
            total_pages = len(pdf.pages)
            get_page = lambda idx: pdf.pages[idx]
            close_doc = lambda: pdf.close()

        print(f"Total pages to parse ({profile.display_name}): {total_pages}")

        try:
            for page_idx in range(total_pages):
                page_num = page_idx + 1
                if self.verbose or page_num % 10 == 0 or page_num == total_pages:
                    print(f"Processing page {page_num}/{total_pages}...")

                page = get_page(page_idx)
                words = self._extract_words(page, engine)

                # Derive column bounds from the page-1 header row (once, reused for all pages)
                if page_idx == 0 and profile.adaptive_bounds:
                    profile = self._apply_adaptive_bounds(words, profile)

                min_y = profile.page1_min_y if page_idx == 0 else profile.pageN_min_y
                table_words = [w for w in words if min_y <= w.top <= profile.max_y]

                # Group words into horizontal lines by Y-coordinate using profile.line_group_tolerance
                grouped_lines = {}
                for w in table_words:
                    top_val = w.top
                    found = False
                    for gk in grouped_lines.keys():
                        if abs(gk - top_val) < profile.line_group_tolerance:
                            grouped_lines[gk].append(w)
                            found = True
                            break
                    if not found:
                        grouped_lines[top_val] = [w]

                sorted_ys = sorted(grouped_lines.keys())

                # Filter out summary/footer lines and column headers with word boundary matching
                valid_ys = []
                for gk in sorted_ys:
                    line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)
                    line_text = " ".join(w.text for w in line_words)

                    # Stop processing page if footer keyword found
                    if any(_matches_word_keyword(kw, line_text) for kw in profile.footer_keywords):
                        break
                    # Skip header lines
                    if any(_matches_word_keyword(kw, line_text) for kw in profile.header_keywords):
                        continue

                    valid_ys.append(gk)

                # ── Anchor & Block Segmentation ──

                if profile.date_type == "single":
                    date_ys = []
                    for gk in valid_ys:
                        line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)
                        dt_text = " ".join(w.text for w in line_words if profile.date_x_range[0] <= w.x0 <= profile.date_x_range[1]).strip()
                        if profile.date_pattern.match(dt_text):
                            date_ys.append((gk, dt_text))

                    # Universal cross-page narration overflow check
                    self._handle_cross_page_continuation(raw_txs, valid_ys, date_ys, grouped_lines, profile)

                    if not date_ys:
                        continue

                    # Hoist column bound lookups outside the word iteration loop
                    val_date_bounds = profile.col_bounds.get('val_date')
                    narration_bounds = profile.col_bounds.get('narration')
                    chq_ref_bounds = profile.col_bounds.get('chq_ref')
                    wx0, wx1 = profile.col_bounds['withdrawal']
                    dx0, dx1 = profile.col_bounds['deposit']
                    bx0, bx1 = profile.col_bounds['balance']

                    for i, (dy, dt_str) in enumerate(date_ys):
                        block_end = date_ys[i + 1][0] if i + 1 < len(date_ys) else (valid_ys[-1] + 1)
                        block_ys = [gk for gk in valid_ys if dy <= gk < block_end]

                        narration_parts = []
                        val_date_text = dt_str
                        chq_ref_parts = []
                        debit_val = None
                        credit_val = None
                        balance_val = None

                        for gk in block_ys:
                            line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)

                            # Value date (if separate column)
                            if val_date_bounds:
                                vx0, vx1 = val_date_bounds
                                vd_text = " ".join(w.text for w in line_words if vx0 <= w.x0 <= vx1).strip()
                                if profile.date_pattern.match(vd_text):
                                    val_date_text = vd_text

                            # Narration
                            if narration_bounds:
                                nx0, nx1 = narration_bounds
                                n_words = [w.text for w in line_words if nx0 <= w.x0 < nx1]
                                if n_words:
                                    narration_parts.append(" ".join(n_words))

                            # Chq/Ref
                            if chq_ref_bounds:
                                rx0, rx1 = chq_ref_bounds
                                if rx0 < rx1:
                                    ref_words = [w.text for w in line_words if rx0 <= w.x0 < rx1 and w.text != '-']
                                    if ref_words:
                                        chq_ref_parts.extend(ref_words)

                            # Amounts: withdrawal, deposit, balance
                            for w in line_words:
                                val = self.clean_amount(w.text)
                                if val is None:
                                    continue

                                if wx0 <= w.x0 < wx1:
                                    debit_val = val
                                elif dx0 <= w.x0 < dx1:
                                    credit_val = val
                                elif bx0 <= w.x0 <= bx1:
                                    balance_val = val

                        if dt_str and balance_val is not None:
                            raw_txs.append({
                                'date': dt_str,
                                'narration': " ".join(narration_parts).strip(),
                                'chq_ref': " ".join(chq_ref_parts).strip(),
                                'val_date': val_date_text,
                                'withdrawal': debit_val,
                                'deposit': credit_val,
                                'balance': balance_val,
                                'page': page_num,
                            })

                elif profile.date_type == "split_3":
                    # Tolerant anchor: join all words in the date column, then match
                    # a combined dd Mon yyyy pattern (case-insensitive month, 1-2 digit day).
                    date_ys = []
                    for gk in valid_ys:
                        line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)
                        joined = " ".join(w.text for w in line_words if profile.date_x_range[0] <= w.x0 <= profile.date_x_range[1]).strip()
                        m = SPLIT_DATE_PATTERN.match(joined)
                        if m:
                            day_text = m.group(1).zfill(2)
                            month_text = m.group(2).title()
                            year_text = m.group(3)
                            date_ys.append((gk, f"{day_text} {month_text} {year_text}"))

                    # Universal cross-page narration overflow check
                    self._handle_cross_page_continuation(raw_txs, valid_ys, date_ys, grouped_lines, profile)

                    if not date_ys:
                        continue

                    # Hoist column bound lookups
                    narration_bounds = profile.col_bounds.get('narration')
                    wx0, wx1 = profile.col_bounds['withdrawal']
                    dx0, dx1 = profile.col_bounds['deposit']
                    bx0, bx1 = profile.col_bounds['balance']

                    for i, (dy, date_text) in enumerate(date_ys):
                        block_end = date_ys[i + 1][0] if i + 1 < len(date_ys) else (valid_ys[-1] + 1)
                        block_ys = [gk for gk in valid_ys if dy <= gk < block_end]

                        narration_parts = []
                        debit_val = None
                        credit_val = None
                        balance_val = None

                        for gk in block_ys:
                            line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)

                            # Narration
                            if narration_bounds:
                                nx0, nx1 = narration_bounds
                                n_words = [w.text for w in line_words if nx0 <= w.x0 <= nx1]
                                if n_words:
                                    narration_parts.append(" ".join(n_words))

                            # Check configurable dash indicators if configured
                            debit_indicator = False
                            if profile.debit_dash_x_range:
                                x_min, x_max = profile.debit_dash_x_range
                                debit_indicator = any(x_min <= w.x0 <= x_max and w.text == '-' for w in line_words)

                            credit_indicator = False
                            if profile.credit_dash_x_range:
                                x_min, x_max = profile.credit_dash_x_range
                                credit_indicator = any(x_min <= w.x0 <= x_max and w.text == '-' for w in line_words)

                            # Guarded assignment: narration-overflow lines carry no
                            # amounts, so a None result must NOT clobber values already
                            # captured from the anchor row.
                            if not debit_indicator:
                                val = self._first_amount_in_range(line_words, wx0, wx1)
                                if val is not None:
                                    debit_val = val

                            if not credit_indicator:
                                val = self._first_amount_in_range(line_words, dx0, dx1)
                                if val is not None:
                                    credit_val = val

                            val = self._first_amount_in_range(line_words, bx0, bx1)
                            if val is not None:
                                balance_val = val

                        if date_text and balance_val is not None:
                            # Balance-chain inference/repair: the running balance delta
                            # mathematically determines withdrawal vs deposit. Parsed
                            # values are kept only when they agree with the chain.
                            debit_val, credit_val = self._infer_split_amounts(
                                raw_txs, balance_val, debit_val, credit_val)
                            raw_txs.append({
                                'date': date_text,
                                'narration': " ".join(narration_parts).strip(),
                                'chq_ref': '',
                                'val_date': date_text,
                                'withdrawal': debit_val,
                                'deposit': credit_val,
                                'balance': balance_val,
                                'page': page_num,
                            })

                elif profile.date_type == "iob":
                    bal_bounds = profile.col_bounds.get('balance')
                    if not bal_bounds:
                        raise ValueError(f"Bank profile '{profile.name}' is missing required 'balance' column bounds")
                    bal_x0 = bal_bounds[0]

                    # Locate date-bearing lines: a genuine IOB row has its date line a
                    # few points ABOVE the amounts line that carries the balance.
                    date_line_ys = set()
                    for gk in valid_ys:
                        line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)
                        dt_text = " ".join(w.text for w in line_words if profile.date_x_range[0] <= w.x0 <= profile.date_x_range[1]).strip()
                        if profile.date_pattern.match(dt_text):
                            date_line_ys.add(gk)

                    # Anchor candidates with validation: only accept balance-column numbers
                    # that have a date line immediately above them. Stray numbers (page
                    # artifacts, summary fragments) no longer create phantom anchors that
                    # split neighbouring transactions' blocks.
                    bal_ys = []
                    for gk in valid_ys:
                        line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)
                        for w in line_words:
                            if w.x0 >= bal_x0 and w.text not in ('-', '(', ')'):
                                val = self._parse_amount_text(w.text)
                                if val is not None:
                                    if any(0 < gk - d <= 12.0 for d in date_line_ys):
                                        bal_ys.append((gk, val))
                                    break

                    # Cross-page narration overflow: pre-anchor lines at the top of a page
                    # always continue the narration of the previous page's final transaction
                    # (raw_txs[-1]), regardless of within-page chronological direction.
                    anchor_date_ys = [(by, "") for by, _ in bal_ys]
                    self._handle_cross_page_continuation(raw_txs, valid_ys, anchor_date_ys, grouped_lines, profile)

                    narration_bounds = profile.col_bounds.get('narration')
                    wx0, wx1 = profile.col_bounds['withdrawal']
                    dx0, dx1 = profile.col_bounds['deposit']

                    for i, (by, bal_val) in enumerate(bal_ys):
                        next_by = bal_ys[i + 1][0] if i + 1 < len(bal_ys) else None
                        block_start = by - profile.iob_block_offset_before
                        block_end = (next_by - profile.iob_block_offset_before) if next_by is not None else (valid_ys[-1] + 1)
                        block_ys = [gk for gk in valid_ys if block_start <= gk < block_end]

                        date_text, val_date_text, ref_text = self._extract_iob_dates_and_ref(block_ys, grouped_lines, profile)
                        narration_parts = []
                        debit_val = None
                        credit_val = None

                        for gk in block_ys:
                            line_words = sorted(grouped_lines[gk], key=lambda w: w.x0)

                            if narration_bounds:
                                nx0, nx1 = narration_bounds
                                n_words = [w.text for w in line_words if nx0 <= w.x0 < nx1]
                                if n_words:
                                    narration_parts.append(" ".join(n_words))

                            # Guarded hardened extraction: dashes/none-values return None
                            # and must not clobber values captured from earlier lines.
                            val = self._first_amount_in_range(line_words, wx0, wx1)
                            if val is not None:
                                debit_val = val
                            val = self._first_amount_in_range(line_words, dx0, dx1)
                            if val is not None:
                                credit_val = val

                        if date_text and bal_val is not None:
                            raw_txs.append({
                                'date': date_text,
                                'narration': " ".join(narration_parts).strip(),
                                'chq_ref': ref_text,
                                'val_date': val_date_text,
                                'withdrawal': debit_val,
                                'deposit': credit_val,
                                'balance': bal_val,
                                'page': page_num,
                            })

        finally:
            close_doc()

        # Handle chronological order adjustment if reverse (e.g. IOB)
        if not profile.chronological:
            page_groups = {}
            for tx in raw_txs:
                page_groups.setdefault(tx['page'], []).append(tx)
            result = []
            for p in sorted(page_groups.keys(), reverse=True):
                result.extend(reversed(page_groups[p]))
            raw_txs = result
            # Now in chronological order: validate/repair W/D via the balance chain.
            raw_txs = self._infer_chain_amounts(raw_txs)

        print(f"Extraction complete. Found {len(raw_txs)} raw transactions.")
        return raw_txs

    # ═══════════════════════════════════════════════════════════════════
    #  Post-processing & output (shared by all banks)
    # ═══════════════════════════════════════════════════════════════════

    def categorize(self, narration):
        """Categorizes a transaction based on narration keywords."""
        text = narration.lower()
        if any(w in text for w in ['salary', 'payroll', 'salary credit', 'betterplace']):
            return 'Salary'
        elif any(w in text for w in ['foreign', 'usd', 'eur', 'inw', 'remittance', 'forex']):
            return 'Foreign Exchange'
        elif any(w in text for w in ['upi-']):
            return 'UPI'
        elif any(w in text for w in ['imps-']):
            return 'IMPS'
        elif any(w in text for w in ['neft-']):
            return 'NEFT'
        elif any(w in text for w in ['rtgs-']):
            return 'RTGS'
        elif any(w in text for w in ['atm-', 'atm wdl']):
            return 'ATM Withdrawal'
        elif any(w in text for w in ['card-', 'pos-', 'pos wdl']):
            return 'Card Payment'
        elif any(w in text for w in ['chq-', 'cheque', 'clg-']):
            return 'Cheque'
        elif any(w in text for w in ['interest credit', 'int.coll', 'interest paid']):
            return 'Interest Income'
        elif any(w in text for w in ['refund', 'cashback']):
            return 'Refund/Cashback'
        elif any(w in text for w in ['charge', 'fee', 'gst-', 'tax', 'annual maint']):
            return 'Bank Charges'
        elif any(w in text for w in ['sweep', 'autosweep', 'mod ']):
            return 'Sweep/MOD'
        elif any(w in text for w in ['mutual fund', 'zerodha', 'groww', 'indmoney', 'icici direct']):
            return 'Investment'
        elif any(w in text for w in ['loan', 'emi-']):
            return 'Loan EMI'
        elif any(w in text for w in ['insurance', 'lic ']):
            return 'Insurance'
        else:
            return 'Other Transfer/Spending'

    def validate_and_process(self, transactions):
        """Validates math equations and adds categorizations."""
        processed = []
        validation_warnings = 0

        print("Validating transaction math integrity...")
        for i, tx in enumerate(transactions):
            tx['category'] = self.categorize(tx['narration'])

            if i > 0:
                prev_balance = processed[i - 1]['balance']
                curr_balance = tx['balance']
                w_amt = tx['withdrawal'] or 0.0
                d_amt = tx['deposit'] or 0.0
                expected_balance = prev_balance - w_amt + d_amt

                if abs(expected_balance - curr_balance) > 0.02:
                    validation_warnings += 1
                    if validation_warnings <= 5:
                        print(f"Warning: Math discrepancy at index {i} (Page {tx['page']}):")
                        print(f"  Prev Balance:     {prev_balance}")
                        print(f"  Withdrawal (-):   {w_amt}")
                        print(f"  Deposit (+):      {d_amt}")
                        print(f"  Expected Balance: {expected_balance:.2f}")
                        print(f"  Parsed Balance:   {curr_balance}")
                        print(f"  Narration:        {tx['narration'][:60]}...")

            processed.append(tx)

        if validation_warnings > 0:
            print(f"Math Validation: WARNING: Found {validation_warnings} discrepancies out of {len(transactions)} transactions.")
        else:
            print("Math Validation: SUCCESS: All transaction balances are mathematically consistent!")

        return processed

    def save_csv(self, transactions, output_path):
        print(f"Saving to CSV: {output_path}")
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Date', 'Narration', 'Chq/Ref No', 'Value Date', 'Withdrawal', 'Deposit', 'Balance', 'Category', 'Page'])
            for tx in transactions:
                writer.writerow([
                    tx['date'],
                    tx['narration'],
                    tx['chq_ref'],
                    tx['val_date'],
                    self.format_amount(tx['withdrawal']),
                    self.format_amount(tx['deposit']),
                    self.format_amount(tx['balance']),
                    tx['category'],
                    tx['page'],
                ])
        print(f"CSV file saved successfully: {os.path.abspath(output_path)}")

    def save_excel(self, transactions, output_path):
        if not HAS_OPENPYXL:
            print("Excel export skipped: openpyxl is not installed.")
            return False

        print(f"Saving to Excel: {output_path}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = ['Date', 'Narration', 'Chq/Ref No', 'Value Date', 'Withdrawal', 'Deposit', 'Balance', 'Category', 'Page']
        ws.append(headers)

        for tx in transactions:
            ws.append([
                tx['date'],
                tx['narration'],
                tx['chq_ref'],
                tx['val_date'],
                tx['withdrawal'] if tx['withdrawal'] is not None else "",
                tx['deposit'] if tx['deposit'] is not None else "",
                tx['balance'] if tx['balance'] is not None else "",
                tx['category'],
                tx['page'],
            ])

        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in range(2, len(transactions) + 2):
            for col in [5, 6, 7]:
                cell = ws.cell(row=row, column=col)
                if cell.value != "":
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.column == 2:
                    max_len = min(max(max_len, len(val_str)), 60)
                else:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        wb.save(output_path)
        print(f"Excel file saved successfully: {os.path.abspath(output_path)}")
        return True


def main():
    parser = argparse.ArgumentParser(description="Convert Bank PDF Statement (HDFC/SBI/IOB/Canara/Indian Bank) to CSV/Excel")
    parser.add_argument("pdf_path", help="Path to Bank Statement PDF file")
    parser.add_argument("--csv", help="Output path for CSV file (default: pdf_name_parsed.csv)")
    parser.add_argument("--xlsx", help="Output path for Excel file (default: pdf_name_parsed.xlsx)")
    parser.add_argument("--bank", choices=['hdfc', 'sbi', 'iob', 'canara', 'indianbank'], help="Bank type (auto-detected by default)")
    parser.add_argument("--verbose", action="store_true", help="Enable verbose log printing")

    args = parser.parse_args()

    if not HAS_FITZ and not HAS_PDFPLUMBER:
        print("Error: Neither 'pymupdf' nor 'pdfplumber' is installed. Please install at least one (e.g. 'pip install pymupdf').", file=sys.stderr)
        sys.exit(1)

    pdf_base = os.path.splitext(args.pdf_path)[0]
    csv_path = args.csv or f"{pdf_base}_parsed.csv"
    xlsx_path = args.xlsx or f"{pdf_base}_parsed.xlsx"

    try:
        start_time = datetime.now()
        parser_inst = BankStatementParser(args.pdf_path, verbose=args.verbose, bank=args.bank)

        raw_txs = parser_inst.parse()
        processed_txs = parser_inst.validate_and_process(raw_txs)
        parser_inst.save_csv(processed_txs, csv_path)

        if HAS_OPENPYXL:
            parser_inst.save_excel(processed_txs, xlsx_path)
        else:
            print("Note: Install 'openpyxl' (pip install openpyxl) to enable formatting Excel .xlsx output.")

        end_time = datetime.now()
        elapsed = (end_time - start_time).total_seconds()

        total_wd = sum(tx['withdrawal'] or 0.0 for tx in processed_txs)
        total_dp = sum(tx['deposit'] or 0.0 for tx in processed_txs)

        print("\n" + "=" * 40)
        print("Parsing Summary")
        print("=" * 40)
        print(f"Processed:          {len(processed_txs)} transactions")
        print(f"Total Withdrawals:  INR {total_wd:,.2f}")
        print(f"Total Deposits:     INR {total_dp:,.2f}")
        print(f"Net Change:         INR {(total_dp - total_wd):,.2f}")
        if processed_txs:
            first_bal = processed_txs[0]['balance']
            first_dep = processed_txs[0]['deposit'] or 0
            first_wd = processed_txs[0]['withdrawal'] or 0
            print(f"Starting Balance:   INR {first_bal - first_dep + first_wd:,.2f}")
            print(f"Ending Balance:     INR {processed_txs[-1]['balance']:,.2f}")
        print(f"Time Elapsed:       {elapsed:.1f} seconds")
        print("=" * 40)

    except Exception as e:
        print(f"Error during parsing: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
